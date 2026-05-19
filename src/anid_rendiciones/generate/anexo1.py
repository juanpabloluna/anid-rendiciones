"""Generador del Anexo N°1 (Formulario de Rendición de Fondos).

El Anexo 1 oficial tiene tres hojas:
  - "Listas" — dropdowns (no se modifica).
  - "Resumen Anexo 1" — cabecera del proyecto y totales por ítem.
  - "Detalle Gastos" — una fila por comprobante (13 columnas).

Esta función carga la plantilla oficial (.xls original convertido a .xlsx)
y la puebla con datos de la rendición. Devuelve los bytes del XLSX listo
para descargar.
"""

from __future__ import annotations

import io
import shutil
import tempfile
from pathlib import Path

import openpyxl
import xlrd
from openpyxl.utils import get_column_letter

from ..schema import Gasto, Item, Proyecto

# Mapeo de columnas en la hoja "Detalle Gastos" (B-indexed, header en fila 8 1-indexed)
COLS_DETALLE = {
    "n_correlativo": "C",
    "item": "D",
    "subitem": "E",
    "rut_beneficiario": "F",
    "nombre_beneficiario": "G",
    "detalle_gasto": "H",
    "tipo_documento": "I",
    "n_documento": "J",
    "fecha_documento": "K",
    "monto_total": "L",
    "monto_rendido": "M",
    "porcentaje_rendido": "N",
    "justificacion": "O",
}
FILA_PRIMER_GASTO = 10  # 1-indexed; el header está en fila 8

# Mapeo de celdas en la hoja "Resumen Anexo 1"
CELDAS_RESUMEN = {
    "fecha_presentacion": "H3",
    "codigo_proyecto": "H9",
    "institucion": "C11",
    "rut_ip": "H11",
    "facultad": "C14",
    "tipo_concurso_etapa": "H14",
    "n_rendicion": "C17",
    "periodo_desde": "G16",
    "periodo_hasta": "H16",
    "monto_transferido": "H33",
}


def _convertir_xls_a_xlsx(xls_path: Path) -> Path:
    """openpyxl no lee .xls. Convertimos a .xlsx usando xlrd + openpyxl manual."""
    workbook_xls = xlrd.open_workbook(str(xls_path), formatting_info=False)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet in workbook_xls.sheets():
        ws = wb.create_sheet(title=sheet.name)
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                if val == "":
                    continue
                ws.cell(row=r + 1, column=c + 1, value=val)

    tmp = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    wb.save(tmp)
    return tmp


def _cargar_plantilla(template_path: Path) -> openpyxl.Workbook:
    if template_path.suffix.lower() == ".xls":
        xlsx_path = _convertir_xls_a_xlsx(template_path)
        return openpyxl.load_workbook(xlsx_path)
    return openpyxl.load_workbook(template_path)


def _escribir_resumen(ws, proyecto: Proyecto, monto_gastado: int) -> None:
    from datetime import date

    ws[CELDAS_RESUMEN["fecha_presentacion"]] = date.today().isoformat()
    ws[CELDAS_RESUMEN["codigo_proyecto"]] = proyecto.codigo
    ws[CELDAS_RESUMEN["institucion"]] = proyecto.institucion_patrocinante
    if proyecto.rut_ip:
        ws[CELDAS_RESUMEN["rut_ip"]] = proyecto.rut_ip
    if proyecto.facultad:
        ws[CELDAS_RESUMEN["facultad"]] = proyecto.facultad
    ws[CELDAS_RESUMEN["tipo_concurso_etapa"]] = f"FONDECYT {proyecto.concurso} {proyecto.anio_concurso} — Etapa {proyecto.etapa}"
    ws[CELDAS_RESUMEN["n_rendicion"]] = proyecto.n_rendicion
    ws[CELDAS_RESUMEN["periodo_desde"]] = proyecto.fecha_inicio_etapa.isoformat()
    ws[CELDAS_RESUMEN["periodo_hasta"]] = proyecto.fecha_fin_etapa.isoformat()
    ws[CELDAS_RESUMEN["monto_transferido"]] = proyecto.monto_transferido


def _escribir_detalle(ws, gastos: list[Gasto]) -> None:
    # La plantilla viene con los correlativos 1..50 prepoblados; limpiar
    # las filas que no usaremos para evitar números huérfanos sin datos.
    for fila in range(FILA_PRIMER_GASTO, FILA_PRIMER_GASTO + 60):
        for col in COLS_DETALLE.values():
            ws[f"{col}{fila}"] = None

    for i, g in enumerate(gastos):
        fila = FILA_PRIMER_GASTO + i
        ws[f"{COLS_DETALLE['n_correlativo']}{fila}"] = i + 1
        if g.item:
            ws[f"{COLS_DETALLE['item']}{fila}"] = g.item.value
        if g.subitem:
            ws[f"{COLS_DETALLE['subitem']}{fila}"] = g.subitem
        if g.rut_beneficiario:
            ws[f"{COLS_DETALLE['rut_beneficiario']}{fila}"] = g.rut_beneficiario
        if g.nombre_beneficiario:
            ws[f"{COLS_DETALLE['nombre_beneficiario']}{fila}"] = g.nombre_beneficiario
        ws[f"{COLS_DETALLE['detalle_gasto']}{fila}"] = g.detalle_gasto
        if g.tipo_documento:
            ws[f"{COLS_DETALLE['tipo_documento']}{fila}"] = g.tipo_documento.value
        if g.n_documento:
            ws[f"{COLS_DETALLE['n_documento']}{fila}"] = g.n_documento
        if g.fecha_documento:
            ws[f"{COLS_DETALLE['fecha_documento']}{fila}"] = g.fecha_documento.isoformat()
        ws[f"{COLS_DETALLE['monto_total']}{fila}"] = g.monto_total
        ws[f"{COLS_DETALLE['monto_rendido']}{fila}"] = g.monto_rendido
        ws[f"{COLS_DETALLE['porcentaje_rendido']}{fila}"] = g.porcentaje_rendido
        if g.justificacion:
            ws[f"{COLS_DETALLE['justificacion']}{fila}"] = g.justificacion


def ordenar_gastos_para_anexo1(gastos: list[Gasto]) -> list[Gasto]:
    """Orden requerido por ANID: por ítem (orden del catálogo), luego por fecha."""
    orden_items = {
        Item.PERSONAL: 0,
        Item.EQUIPAMIENTO: 1,
        Item.INFRAESTRUCTURA: 2,
        Item.OPERACION: 3,
        Item.INDIRECTOS: 4,
    }

    def key(g: Gasto):
        item_orden = orden_items.get(g.item, 99) if g.item else 99
        fecha = g.fecha_documento or g.fecha_documento  # None ordena al final
        from datetime import date

        return (item_orden, g.subitem or "", fecha or date.max)

    return sorted(gastos, key=key)


def generar_anexo1(
    proyecto: Proyecto,
    gastos: list[Gasto],
    plantilla_path: Path,
) -> bytes:
    """Genera el Anexo 1 poblado y devuelve los bytes del XLSX."""
    gastos_ordenados = ordenar_gastos_para_anexo1(gastos)

    wb = _cargar_plantilla(plantilla_path)

    # Ubicar las hojas — toleramos cambios de nombre/case
    nombres = {sn.lower().strip(): sn for sn in wb.sheetnames}
    resumen_name = next((nombres[k] for k in nombres if "resumen" in k), None)
    detalle_name = next((nombres[k] for k in nombres if "detalle" in k), None)

    if resumen_name:
        monto_gastado = sum(g.monto_rendido for g in gastos_ordenados)
        _escribir_resumen(wb[resumen_name], proyecto, monto_gastado)

    if detalle_name:
        _escribir_detalle(wb[detalle_name], gastos_ordenados)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
